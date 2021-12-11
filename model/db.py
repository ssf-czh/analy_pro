import base64

from model.schema import *
from typing import List
from template import template_file_base64
import openpyxl
import os

init_params = InitialParameters()
main_fittings: List[MainFitting] = list()
pump2_fittings: List[Pump2Fitting] = list()
pump3_fittings: List[Pump3Fitting] = list()
wet_bulb_fittings: List[WetBulbFitting] = list()
p4_fittings: List[P4Fitting] = list()
fitting_coefficients = FittingCoefficients()
optimize_result: List[OptimizeResult] = list()


def load(path: str):
    global init_params, main_fittings, pump2_fittings, pump3_fittings, wet_bulb_fittings, p4_fittings, \
        fitting_coefficients, optimize_result
    if not os.path.exists(path):
        save_default(path)
    wb = openpyxl.load_workbook(path, read_only=True)
    for sheet in wb:
        if sheet.title == "参数初始值设定":
            init_params.q = float(sheet["B4"].value)
            init_params.n = int(sheet["B5"].value)
            init_params.efficiency_range = (float(sheet["B6"].value), float(sheet["C6"].value))
            init_params.t3_min = float(sheet["B7"].value)
            init_params.g = float(sheet["B9"].value)
            init_params.h = float(sheet["B10"].value)
            init_params.p2 = float(sheet["B11"].value)
            init_params.delta_t1_range = (float(sheet["B24"].value), float(sheet["C24"].value))
            init_params.delta_t2_range = (float(sheet["B25"].value), float(sheet["C25"].value))
        elif sheet.title == "主机参数拟合":
            main_fittings = list()
            i = 3
            while sheet.cell(i, 1).value is not None:
                entry = MainFitting()
                entry.load_percentage = float(sheet.cell(i, 1).value)
                entry.q = float(sheet.cell(i, 2).value)
                entry.p1 = float(sheet.cell(i, 3).value)
                entry.t1 = float(sheet.cell(i, 4).value)
                entry.t2 = float(sheet.cell(i, 5).value)
                entry.t3 = float(sheet.cell(i, 6).value)
                entry.t4 = float(sheet.cell(i, 7).value)
                entry.cop = float(sheet.cell(i, 8).value)
                main_fittings.append(entry)
                i += 1
        elif sheet.title == "水泵性能参数拟合":
            pump2_fittings = list()
            for i in range(5):
                entry = Pump2Fitting()
                entry.g2 = float(sheet.cell(4, 3 + i * 2).value)
                entry.p2 = float(sheet.cell(4, 4 + i * 2).value)
                pump2_fittings.append(entry)

            pump3_fittings = list()
            for i in range(5):
                entry = Pump3Fitting()
                entry.g3 = float(sheet.cell(11, 3 + i * 2).value)
                entry.p3 = float(sheet.cell(11, 4 + i * 2).value)
                pump3_fittings.append(entry)
        elif sheet.title == "冷却塔拟合":
            i = 0
            wet_bulb_fittings = list()
            while sheet.cell(3 + i, 1).value is not None:
                entry = WetBulbFitting()
                entry.temp = float(sheet.cell(3 + i, 1).value)
                entry.amplitude = float(sheet.cell(3 + i, 2).value)
                wet_bulb_fittings.append(entry)
                i += 1

            i = 0
            p4_fittings = list()
            while sheet.cell(3 + i, 4).value is not None:
                entry = P4Fitting()
                entry.g = float(sheet.cell(3 + i, 4).value)
                entry.p4 = float(sheet.cell(3 + i, 5).value)
                p4_fittings.append(entry)
                i += 1
        elif sheet.title == "拟合系数表":
            fitting_coefficients = FittingCoefficients()
            for i in range(12):
                val = 0.0
                if sheet.cell(3, 2 + i).value is not None:
                    val = float(sheet.cell(3, 2 + i).value)
                fitting_coefficients.b.append(val)
            for i in range(12):
                val = 0.0
                if sheet.cell(5, 2 + i).value is not None:
                    val = float(sheet.cell(3, 2 + i).value)
                fitting_coefficients.b.append(val)
            for i in range(3):
                val = 0.0
                if sheet.cell(8, 2 + i).value is not None:
                    val = float(sheet.cell(8, 2 + i).value)
                fitting_coefficients.a.append(val)
            for i in range(3):
                val = 0.0
                if sheet.cell(11, 2 + i).value is not None:
                    val = float(sheet.cell(11, 2 + i).value)
                fitting_coefficients.c.append(val)
            for i in range(3):
                val = 0.0
                if sheet.cell(14, 2 + i).value is not None:
                    val = float(sheet.cell(14, 2 + i).value)
                fitting_coefficients.d.append(val)
            for i in range(4):
                val = 0.0
                if sheet.cell(16, 2 + i).value is not None:
                    val = float(sheet.cell(16, 2 + i).value)
                fitting_coefficients.e.append(val)
        elif sheet.title == "优化计算结果":
            optimize_result = list()
            i = 0
            while sheet.cell(2 + i, 1).value is not None:
                entry = OptimizeResult()
                entry.q = float(sheet.cell(2 + i, 1).value)
                entry.ts = float(sheet.cell(2 + i, 2).value)
                optimize_result.append(entry)
                i += 1
    wb.close()


def save_default(path: str):
    with open(path, "wb") as f:
        data = base64.decodebytes(bytes(template_file_base64, "utf-8"))
        f.write(data)
        f.close()


def save(path: str):
    if not os.path.exists(path):
        save_default(path)
    wb = openpyxl.load_workbook(path, read_only=False)
    for sheet in wb:
        if sheet.title == "参数初始值设定":
            sheet["B4"].value = str(init_params.q)
            sheet["B5"].value = str(init_params.n)
            sheet["B6"].value = str(init_params.efficiency_range[0])
            sheet["C6"].value = str(init_params.efficiency_range[1])
            sheet["B7"].value = str(init_params.t3_min)
            sheet["B9"].value = str(init_params.g)
            sheet["B10"].value = str(init_params.h)
            sheet["B11"].value = str(init_params.p2)
            sheet["B24"].value = str(init_params.delta_t1_range[0])
            sheet["C24"].value = str(init_params.delta_t1_range[1])
            sheet["B25"].value = str(init_params.delta_t2_range[0])
            sheet["C25"].value = str(init_params.delta_t2_range[1])
        elif sheet.title == "主机参数拟合":
            i = 3
            for entry in main_fittings:
                sheet.cell(i, 1).value = str(entry.q / 2813.0)
                sheet.cell(i, 2).value = str(entry.q)
                sheet.cell(i, 3).value = str(entry.p1)
                sheet.cell(i, 4).value = str(entry.t1)
                sheet.cell(i, 5).value = str(entry.t2)
                sheet.cell(i, 6).value = str(entry.t3)
                sheet.cell(i, 7).value = str(entry.t4)
                sheet.cell(i, 8).value = str(entry.cop)
                i += 1
            while sheet.cell(i, 1).value is not None:
                sheet.delete_rows(i)
        elif sheet.title == "水泵性能参数拟合":
            for i in range(5):
                sheet.cell(4, 3 + i * 2).value = str(pump2_fittings[i].g2)
                sheet.cell(4, 4 + i * 2).value = str(pump2_fittings[i].p2)
            for i in range(5):
                sheet.cell(11, 3 + i * 2).value = str(pump3_fittings[i].g3)
                sheet.cell(11, 4 + i * 2).value = str(pump3_fittings[i].p3)
        elif sheet.title == "冷却塔拟合":
            i = 0
            # wet_bulb_fittings = list()
            for entry in wet_bulb_fittings:
                sheet.cell(3 + i, 1).value = str(entry.temp)
                sheet.cell(3 + i, 2).value = str(entry.amplitude)
                i += 1
            while sheet.cell(3 + i, 1).value is not None:
                sheet.cell(3 + i, 1).value = None
                sheet.cell(3 + i, 2).value = None
                i += 1

            i = 0
            for entry in p4_fittings:
                sheet.cell(3 + i, 4).value = str(entry.g)
                sheet.cell(3 + i, 5).value = str(entry.p4)
                i += 1
            while sheet.cell(3 + i, 4).value is not None:
                sheet.cell(3 + i, 4).value = None
                sheet.cell(3 + i, 5).value = None
                i += 1
        elif sheet.title == "拟合系数表":
            for i in range(12):
                sheet.cell(3, 2 + i).value = str(fitting_coefficients.b[i])
            for i in range(12):
                sheet.cell(5, 2 + i).value = str(fitting_coefficients.b[i + 12])
            for i in range(3):
                sheet.cell(8, 2 + i).value = str(fitting_coefficients.a[i])
            for i in range(3):
                sheet.cell(11, 2 + i).value = str(fitting_coefficients.c[i])
            for i in range(3):
                sheet.cell(14, 2 + i).value = str(fitting_coefficients.d[i])
            for i in range(4):
                sheet.cell(16, 2 + i).value = str(fitting_coefficients.e[i])
        elif sheet.title == "优化计算结果":
            i = 0
            for entry in optimize_result:
                sheet.cell(2 + i, 1).value = str(entry.q)
                sheet.cell(2 + i, 2).value = str(entry.ts)
                sheet.cell(2 + i, 4).value = str(entry.load_percentage)
                sheet.cell(2 + i, 5).value = str(entry.t1)
                sheet.cell(2 + i, 6).value = str(entry.t2)
                sheet.cell(2 + i, 7).value = str(entry.t3)
                sheet.cell(2 + i, 8).value = str(entry.t4)
                sheet.cell(2 + i, 9).value = str(entry.delta_t)
                sheet.cell(2 + i, 10).value = str(entry.p1)
                sheet.cell(2 + i, 11).value = str(entry.p2)
                sheet.cell(2 + i, 12).value = str(entry.p3)
                sheet.cell(2 + i, 13).value = str(entry.p4)
                sheet.cell(2 + i, 14).value = str(entry.p)
                sheet.cell(2 + i, 15).value = str(entry.cop)
                sheet.cell(2 + i, 16).value = str(entry.n)
                i += 1
            while sheet.cell(2 + i, 1).value is not None:
                for j in range(1, 17):
                    sheet.cell(2 + i, j).value = None
                i += 1
    wb.save(path)
    wb.close()


"""
Test
if __name__ == '__main__':
    fitting_coefficients.b = [float(x) for x in range(24)]
    fitting_coefficients.a = [0.0, 1.0, 2.0]
    fitting_coefficients.c = [0.0, 1.0, 2.0]
    fitting_coefficients.d = [0.0, 1.0, 2.0]
    fitting_coefficients.e = [0.0, 1.0, 2.0, 3.0]
    for i in range(10):
        main_fittings.append(MainFitting())
    for i in range(5):
        pump2_fittings.append(Pump2Fitting())
        pump3_fittings.append(Pump3Fitting())
    for i in range(2):
        wet_bulb_fittings.append(WetBulbFitting())
    p4_fittings.append(P4Fitting())
    entry = OptimizeResult()
    entry.Q = 1145.0
    entry.ts = 233.0
    optimize_result.append(entry)
    save("testtest.xlsx")
"""
